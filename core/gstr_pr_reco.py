import pandas as pd
import numpy as np
import os
import glob
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from shutil import copyfile
import datetime
import warnings
import re
from difflib import SequenceMatcher , get_close_matches

from .CONSTANTS import * #added this line


####ADDED THIS LINE OF CODE ##################
# if not os.path.exists('/upload'):
#     os.makedirs('/upload')
############## END ########################

warnings.filterwarnings('ignore')

def getgstcheck(number):


    """
    This function getgstcheck will give the last digit of the gst number

    Output will be given as the last digit which should be as per the given 14 digit number

    :param number: This is the only argument that needs to be given. It is a mandatory Argument.The argument must beat least 14 digit long

    :type gst_no: This parameter must be a string and must be at least digit long

    :return : The function will return the correct last digit of the given gst number

    :raises: There are two errors that will be raised by the Function 
            1. Type Error: If the parameter entered is not a string, then this error is raised
            2. Exception: If the parameter entered is not at least 14 digit long, then Exception is raised

    :see also: To know how the GST Ceck sum is calculated , see the alogorithm behind the last digit



    """

    charlist=[char for char in number.upper()]

    a=1
    cumhash=[]

    if not type(number) is str:
        raise TypeError("Only strings are allowed")


    elif len(str((number)))<14:
        
        
        print ("Please ensure that the input is at least 14 digit long")
        
        pass

    else:
        
        pass
        


    for i in charlist[0:14:1]:
        
        if a % 2==0:
            multiplier=2
        else:
            multiplier=1

        if i.isdigit():
            intvalue=int(i)
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            
        else:
            intvalue=ord(i)-55   
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            

        a=a+1

        cumhash.append(hash)

    hashsum=(sum(cumhash))

    remain=hashsum%36

    checksum=36-remain

    if checksum<10:
        finalchk=str(checksum)
    
    elif checksum==36:
        finalchk=str(0)
    else:
        finalchk=chr(checksum+55)
   
    
    return (finalchk)


def gstchecksum(gst_no):


    """
    This function gstchecksum will check the last digit of the gst number and return whether the Check sum matches or not

    Output will be given as "Check Sum MATCH" or "Check Sum MISMATCH"

    :param gst_no: This is the only argument that needs to be given. It is a mandatory Argument.The argument must be 15 digit long

    :type gst_no: This parameter must be a string and must be 15 digit long

    :return : The function will return only one of two values a) Check Sum MATCH & b) Check Sum MISMATCH

    :raises: There are two errors that will be raised by the Function 
            1. Type Error: If the parameter entered is not a string, then this error is raised
            2. Exception: If the parameter entered is not 15 digit long, then Exception is raised

    :see also: To know how the GST Ceck sum is calculated , see the alogorithm behind the last digit



    """
   
    lastchr=str(gst_no[len(gst_no)-1])
   
        
    if getgstcheck(gst_no)==lastchr:
        result="Check Sum MATCH"
    else:
        result="Check Sum MISMATCH"
    
    return (result)
    

def gstinvcheck(a):

    """
    This function will check whether the invoice number entered is correct or not.

    As per GST rules, the Invoice number must be maximum 15 digit long

    :param a: this must be the GST  Invoice number 

    :param a type: The Type of parameter must be a string. However, in the functionit is converting any parameter into a string through str() method

    :return : it return one of 2 output  a) Invoice Number Valid or b) Invoice Number Invalid



    """

    try:
        length=len(str(a))
    except:
        length=0    


    if length<=16:
        status="Invoice Number Valid"
    else:
        status="Invoice Number Invalid"
        
    return(status)


def extract_pan(gst_no):



    """
    This function will extract the PAN number from the provided GST_No

    :param gst_no: This function requires only one parameter. ie the GST No

    :param gst_no type: The type of the parameter must be a string

    :return :The function will return a string which is the PAN Number

    :SeeAlso : The PAN number is the 3rd Character to 12th Character of the GST Number



    """
    try:
        if not type(gst_no) is str:
            raise TypeError("Only strings are allowed")

        else:
            pass
       
    except:
        pass
        
    try:
        if len(gst_no)<15:
            raise Exception("Please ensure that the input is 15 digit long")

        else:
            pass
            
    except:
        pass

    

    try:
        pan_num=gst_no[2:12:1]
    except:
        pan_num=gst_no
    
    return(pan_num)



def download(pth=os.getcwd()):



    """
    This is a function to download the GSTR2A & ITR format and also instructions to use this utility


    :param pth: This takes a single argument which is a pathn which the user wants to store the Format files

    :param Type: This parameter is a optional argument

                In case the parameter is not provided, the current working directory is taken as the Pth and the fomrat os downloaded in that folder


    :return writer: This function will return a excel file which has a format for the reconciliation of the GSTR2A and the ITR

    This is a dependednt function for the next main function reco_itr_2a. 

    There are mandatory columns and it has to be ensured that the names of the Mandatory columns are same as in the format

    There is no requirment for the sequence of the columns to be same as the Format

    The excel file in which the data is kept can hae multiple sheets , but the nme of the sheet should be same as in the format

    For more details, refer the Sheet "Important_Checklist" downloaded in the format


    """



    import pandas as pd
    import numpy as np
    import openpyxl
    # from CONSTANTS import * #commented this
    


    fullpath1 = pth + "\\" + "Formats.xlsx"
    print(f"The path selected is {fullpath1}")

    # writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True}) #BOB COMMENTED THIS
    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}}) #BOB ADDED THIS

    dict1 = {PR_VENDOR_GSTREG_NO: ["Mandatory"], PR_VENDOR_NAME: ["Optional"], PR_INVOICE_NUMBER: ["Mandatory"],
             PR_INVOICE_DATE_TEXT: ["Mandatory"], PR_TOTAL_TAX: ["Mandatory"], PR_IGST: ["Optional"], PR_SGST: ["Optional"],
             PR_CGST: ["Optional"], PR_UTGST: ["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}

    df1 = pd.DataFrame(dict1)
    df1.to_excel(writer, sheet_name=ITR_SHEET_NAME, index=False)

    dict2={R2_VENDOR_GSTREG_NO:["Mandatory"],R2_INVOICE_NUMBER:["Mandatory"],R2_VENDOR_NAME:["Optional"],R2_INVOICE_DATE_TEXT:["Mandatory"],R2_TOTAL_TAX:["Mandatory"],R2_R1_FILING_STATUS:["Mandatory"],R2_ATTRACT_RCM:["Mandatory"],R2_IGST:["Optional"],R2_SGST:["Optional"],R2_CGST:["Optional"],R2_UTGST:["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}
    df2 = pd.DataFrame(dict2)
    df2.to_excel(writer, sheet_name=GSTR2A_SHEET_NAME, index=False)

    dict3={"Things to ensure before Running the Program":[" ","Keep GSTR2A and ITR in different Folders ","In that Folder , there should not be any other excel files",
    "Format of the ITR & GSTR2A can be as per the user ","However, below points to be taken care","The name of the Sheet having the  ITR should be Main_ITR_Format ",
    "The name of the Sheet having the Consolidated GSTR2A should be Main_2A_Format "," ","There are 6 Mandatory columns in  GSTR2A and 4 Mandatory columns in ITR ","The Name to be assigned to these 6 Mandatory columns must be same as in the format",
    "In ITR , Mandatory columns are Vendor_GST_REG , Invoice_Number, Invoice_Date_Text, Total_Tax","Even the upper and Lower case should be same as in the Format",
    "In GSTR2A , Mandatory columns are GSTIN_of_Supplier , Inv_CN_DN_Number_Final, Inv_CN_DN_Date_Text, Total_Tax,Supply_Attract_Reverse_Charge,GSTR_1_5_Filing_Status","Take care of the Upper and Lower case and special Character",
    "The sequence of the Columns is not relevant. User can maintain the Sequence of the columns as per his own convinience","The ITR or GSTR2A file can also have multiple other sheets as per need of user , but relevant data for matching should be in one sheet only .",
     "  "," ","For any issues in running the Code, send your issues to efficientcorporates.info@gmail.com",
    "For more such Automation Videos, Follow YouTube Channel Efficient Corporates"]}
    
    df3 = pd.DataFrame(dict3)
    df3.to_excel(writer, sheet_name=CHECKLIST_SHEET_NAME, index=False)


    # writer.save()   #BOB COMMENTED THIS

    writer.close()

    print(f'The Formats have been saved in below path \n {fullpath1}\n ')

    # return (writer)
    return {
        "writer":writer,
        "fullpath1":fullpath1
    }


def reco_itr_2a(files_itr,files_con2a,tol_limit=100):

    """
    This fucntion is for reconciling the GSTr2A and the ITR

    Please download the Format using the Function download and go through the Important Checklist

    This function takes the 3 parameters. Two are Mandatory and 1 is optional

    :param files_itr: This argument should be the complete path to the ITR file which is as per the format 

                    Please ensure to provide the compelte filepath of ITR till the extension

    :param files_con2a : This is the argument for the complete filepath of the GSTR2A file.
                        
                        Please ensure to gve the complete file path till the extension

    :param tol_limit : This is also next important parameter. This is the Tolerance limit.

                        If a invoice is booked with Tax of Rs 12,300 , but the same invoice is given in GSTR2A as Rs 12450.

                        Now, there is a difference of Rs 150. Now , if the tolerance limit is kept as 100, then this case will be considered NOT MATCHING

                        But, if the tolerance limit is kept as 200, then this case will be considered as a match

                        Use can provide the Tolerance limit value based on the size of the client


                        If no parameter is provided , then the 100 is taken as the Tolerance limit

    :return output : This function will return 2 files 1) Summary.xlsx and 2) Working.xlsx

                    These 2 files will be stored in the folder where the Combined GSTR2A is stored

                    The Summary fil will ave a snapshot of the matching exercise and will tell the Total cases, matched cases and the unmatched cases

                    The matching is done under 7 different categories

                    a) GST+INV NO + INV Date +Tax Amount >> Complete 3 way match

                    b)GST + INV NO +Tax Amount >> Complete 2 way match

                    c)GST + INV Date +Tax Amount >> Complete 2 way match

                    d)PAN+ INV NO + INV Date +Tax Amount >> Complete 3 way match

                    e)PAN+INV NO  +Tax Amount >> Complete 2 way match

                    f) PAN + INV Date +Tax Amount >> Complete 2 way match

                    g) Fuzzy Look up Match: These are the cases with 90% Invoice Number and 100% PAN Number matching . Just the Invoice Number matches , not the Tax Amount or date


                    Also, the Unmatched cases of ITR will be bifurcated into 3 difefrent buckets 

                    1. Cases whose GST/PAN is not present in GSTR2A (No Scope of Mathing)
                    2. Cases where the GST Number entered in Purchase Register is INVALID
                    3. Cases where the Invoice Number is Invalid
                    4. Other Remaining Unmatched Cases 



    """



    import numpy as np
    import openpyxl

    # from CONSTANTS import *   #commented this


    

    warnings.filterwarnings('ignore')



    print(f'The Consolidated GSTR2A file path is {files_con2a}')
    print(f'The ITR file path is {files_itr}')

    pth = os.path.dirname(str(files_con2a))
    
    fullpath1 = pth + "/" + EXPORT_WORKING_NAME
    
    # writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})
    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_formulas': True}})
    




    fullpath1a = pth + "/" + EXPORT_SUMMARY_NAME

    # writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', options={'strings_to_formulas': True})
    writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_formulas': True}})

    df1 = pd.DataFrame()
    df1.to_excel(writer1, sheet_name=EXPORT_SUMMARY_SHEET, index=False)

    # writer1.save()
    writer1.close()

    fullpath2 = fullpath1a.replace("/", "\\")  # this is a very useful command for defining the correct filepath

    wb = load_workbook(fullpath2)
    ws = wb[EXPORT_SUMMARY_SHEET]

    ws["B2"].value = "SUMMARY OF THE RECONCILIATION OF GSTR2A Vs ITR"
    ws.merge_cells("B2:F2")
    ws["C4"].value = "GSTR2A"
    ws.merge_cells("C4:D4")
    ws["E4"].value = "Purchase Register"
    ws.merge_cells("E4:F4")

    ws["B4"].value = "Particulars"
    ws.merge_cells("B4:B5")
    ws["C5"].value = "Count"
    ws["D5"].value = "Tax Amount"
    ws["E5"].value = "Count"
    ws["F5"].value = "Tax Amount"

    ws["B7"].value = "Total cases in Original Files"
    ws["B8"].value = "Less: No GST Number in Purchase Register "
    ws["B9"].value = "Less: GSTR-1 Not filed cases"
    ws["B10"].value = "Less: GSTR-1 Filed but RCM cases "
    ws["B11"].value = "Less: No Invoice Number in Purchase Register or NO Tax Amount in GSTR2A"



    ws["B12"].value = "Net cases to be Matched"
    ws["B14"].value = "Matched with GST_INVNO_INVDATE_3_WAY"
    ws["B15"].value = "Matched with GST_INVNO_2_WAY"
    ws["B16"].value = "Matched with GST_INVDATE_2_WAY"
    ws["B17"].value = "3 Way Matched with PAN - Diff in Amount Beyond Tolerance Limit"

    ws["B18"].value = "Identified Possible Matches - Fuzzy Logic"

    ws["B20"].value = "Matched with PAN_INVNO_INVDATE_3_WAY"
    ws["B21"].value = "Matched with PAN_INVNO_2_WAY"
    ws["B22"].value = "Matched with PAN_INVDATE_2_WAY"

    ws["B24"].value = "Unmatched Cases"


    ws["B25"].value = "Unmatched Cases -PAN/GST not available in GSTR2A"
    ws["B26"].value = "Unmatched Cases with Invalid GSTIN"


    ws["B27"].value = "Unmatched cases with Invalid Invoice Number"

    ws["B28"].value = "Other Unmatched Cases"

    ws["B30"].value = "Check"

    # setting the tolerance limit for matching in Rupees

    tol_limit = int(tol_limit)

    print(f"The tolerance limit is set to {tol_limit}")

    ws["F1"].value = f"Tolerance Limit was {tol_limit}"

    gstr2a = pd.read_excel(files_con2a, sheet_name=GSTR2A_SHEET_NAME,dtype={R2_INVOICE_NUMBER:str, R2_INVOICE_DATE_TEXT:str, R2_TOTAL_TAX:float})

    try:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a[R2_INVOICE_NUMBER].apply(lambda x: x.lower(str()))
    except:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a[R2_INVOICE_NUMBER]

    try:
        gstr2a['Inv_CN_DN_Date_Text'] = gstr2a[R2_INVOICE_DATE_TEXT].apply(lambda x: x.replace("/",".").replace("-", "."))
    except:
        gstr2a['Inv_CN_DN_Date_Text'] = gstr2a[R2_INVOICE_DATE_TEXT]


    gstr2a['GST_INVNO_INVDATE_3_WAY'] = gstr2a[R2_VENDOR_GSTREG_NO] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + gstr2a[R2_INVOICE_DATE_TEXT]

    gstr2a['GST_INVNO_2_WAY'] = gstr2a[R2_VENDOR_GSTREG_NO] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['GST_INVDATE_2_WAY'] = gstr2a[R2_VENDOR_GSTREG_NO] + "/" + gstr2a["Inv_CN_DN_Date_Text"]

    
    gstr2a['PAN_Number'] = gstr2a[R2_VENDOR_GSTREG_NO].apply(lambda x:extract_pan(x))
    
    # the PAN number matches will be used as possible matches

    gstr2a['PAN_INVNO_INVDATE_3_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + \
                                        gstr2a["Inv_CN_DN_Date_Text"]

    gstr2a['PAN_INVNO_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['PAN_INVDATE_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a["Inv_CN_DN_Date_Text"]


    itr = pd.read_excel(files_itr, sheet_name=ITR_SHEET_NAME,dtype={PR_INVOICE_NUMBER:str, PR_INVOICE_DATE_TEXT:str,PR_TOTAL_TAX:object})

    # itr = pd.read_excel(files_itr, sheet_name="Main_ITR_Format")


    try:
        itr["Invoice_Numberl"] = itr[PR_INVOICE_NUMBER].apply(lambda x: x.lower(str()))
    except:
        itr["Invoice_Numberl"] = itr[PR_INVOICE_NUMBER]

    try:

        itr["Invoice_Date_Text"] = itr[PR_INVOICE_DATE_TEXT].apply(lambda x: x.replace("/",".").replace("-", "."))
    except:
        itr["Invoice_Date_Text"] = itr[PR_INVOICE_NUMBER]


    itr["GST_INVNO_INVDATE_3_WAY"] = itr[PR_VENDOR_GSTREG_NO] + "/" + itr["Invoice_Numberl"] + "/" + itr[
        "Invoice_Date_Text"]

    itr["GST_INVNO_2_WAY"] = itr[PR_VENDOR_GSTREG_NO] + "/" + itr["Invoice_Numberl"]

    itr["GST_INVDATE_2_WAY"] = itr[PR_VENDOR_GSTREG_NO] + "/" + itr["Invoice_Date_Text"]

   
    itr["PAN_Number"] = itr[PR_VENDOR_GSTREG_NO].apply(lambda x:extract_pan(x))
    


    # the PAN number matches will be used as possible matches

    itr["PAN_INVNO_INVDATE_3_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"] + "/" + itr["Invoice_Date_Text"]

    itr["PAN_INVNO_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"]

    itr["PAN_INVDATE_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Date_Text"]

    ws["C7"].value = list(gstr2a.shape)[0]
    ws["D7"].value = sum(gstr2a["Total_Tax"])
    ws["E7"].value = list(itr.shape)[0]
    ws["F7"].value = sum(itr["Total_Tax"])



    #data Cleaning for GSTr2A:

    try:
        
        gstr2a_not_filed=gstr2a[gstr2a[R2_R1_FILING_STATUS] == "N"]
        gstr2a_not_filed["Final_Effcorp _Remarks"]="GSTR-1 Status is Not Filed"

    except:

        gstr2a_not_filed=pd.DataFrame()


    try:
        gstr2a_rcm=gstr2a[(gstr2a[R2_ATTRACT_RCM] == "Y") & (gstr2a[R2_R1_FILING_STATUS] == "Y")]
        gstr2a_rcm["Final_Effcorp _Remarks"]="GSTR-1 Filed Under RCM- No Matching required"
    except:
        gstr2a_rcm=pd.DataFrame()
    
    
    try:
        gstr2a_notax=gstr2a[(gstr2a[R2_TOTAL_TAX] < 1) & (gstr2a[R2_ATTRACT_RCM] != "Y") & (gstr2a[R2_R1_FILING_STATUS] == "Y")]
        gstr2a_notax["Final_Effcorp _Remarks"]="Tax Amount is Zero or less than Re 1"
    except:
        gstr2a_notax=pd.DataFrame()
    
    


    try:
        gstr2a_work=gstr2a[(gstr2a[R2_TOTAL_TAX] >= 1) & (gstr2a[R2_ATTRACT_RCM] != "Y") & (gstr2a[R2_R1_FILING_STATUS] == "Y")]
           
    except:
        
        #in the below blok of code ,w e are using the outer functionaly, which returns a indicator , leftonly, right nly and both , and we will 
        #be using the left only by filtering that out
        
        
        gstr2a_work_3=pd.merge(gstr_2a,gstr2a_not_filed,how='outer',indicator=True)
        mask_3=gstr2a_work_3['_merge']=='left_only'
        
        gstr2a_work_3=gstr2a_work_3[mask_3]
#         print(gstr2a_work_2)
        
        
        gstr2a_work_2=pd.merge(gstr_2a_work_3,gstr2a_rcm,how='outer',indicator=True)
        mask_2=gstr2a_work_2['_merge']=='left_only'
        
        gstr2a_work_2=gstr2a_work_2[mask_2]
#         print(gstr2a_work_2)
        
        
        gstr2a_work_1=pd.merge(gstr_2a_work_2,gstr2a_notax,how='outer',indicator=True)
        mask_1=gstr2a_work_1['_merge']=='left_only'
        
        gstr2a_work_1=gstr2a_work_1[mask_1]
#         print(gstr2a_work_2)


        gstr2a_work=gstr2a_work_1     
        
        
    
    ws["C9"].value = len(gstr2a_not_filed[R2_VENDOR_GSTREG_NO])
    ws["D9"].value = sum(gstr2a_not_filed[R2_TOTAL_TAX])
    
    ws["C10"].value = len(gstr2a_rcm[R2_VENDOR_GSTREG_NO])
    ws["D10"].value = sum(gstr2a_rcm[R2_TOTAL_TAX])

    
    ws["C11"].value = len(gstr2a_notax[R2_VENDOR_GSTREG_NO])
    ws["D11"].value = sum(gstr2a_notax[R2_TOTAL_TAX])

    
    ws["C12"].value = len(gstr2a_work[R2_VENDOR_GSTREG_NO])
    ws["D12"].value = sum(gstr2a_work[R2_TOTAL_TAX])


    #data cleaing  for ITR as of now is only Blank GST Reg No and Blank Invoice Number.
    #So, net case to be matched will equal to Total cases

    mask=itr[PR_VENDOR_GSTREG_NO].isnull()

    try:
        itr_nogst=itr[mask]
        itr_nogst["Final_Effcorp _Remarks"]="NO GST Number in Purchase Register"
        
    except:
        itr_nogst=pd.DataFrame()


    itr_gst=itr[~mask]

    mask2=itr_gst["Invoice_Numberl"].isnull()

    try:
        itr_noinvno=itr_gst[mask2]
        itr_noinvno["Final_Effcorp _Remarks"]="No Invoice Number in Purchase Register"
    except:
        itr_noinvno=pd.DataFrame()

    itr_work=itr_gst[~mask2]


    ws["E8"].value = list(itr_nogst.shape)[0]
    ws["F8"].value = sum(itr_nogst[PR_TOTAL_TAX])

    ws["E11"].value = list(itr_noinvno.shape)[0]
    ws["F11"].value = sum(itr_noinvno[PR_TOTAL_TAX])

    ws["E12"].value = list(itr_work.shape)[0]
    ws["F12"].value = sum(itr_work[PR_TOTAL_TAX])


    # First Cut Matching : Here we will try to do that Matching based on 3 way i.e GST No, Inv No & Inv Date being same in ITR & GSTR2A

    
    gstr2a_pivot = pd.pivot_table(gstr2a_work, values=R2_TOTAL_TAX, index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(itr_work, values=R2_TOTAL_TAX, index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_INVDATE_3_WAY", right_on="GST_INVNO_INVDATE_3_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_3_way_list = compared[mask_1]["GST_INVNO_INVDATE_3_WAY"].values

    mask_1a = gstr2a_work["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a Boolean Array

    mask_1b = itr_work["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a boolean array

    matched_gstr2a_3way = gstr2a_work[mask_1a]
    matched_gstr2a_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"
    matched_gstr2a_3way["Final_Effcorp _Remarks"] = "Exact Matches with Amount- GSTIN 3Way"
    
    matched_itr_3way = itr_work[mask_1b]
    matched_itr_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"
    matched_itr_3way["Final_Effcorp _Remarks"] = "Exact Matches with Amount- GSTIN 3Way"

    ws["C14"].value = len(matched_gstr2a_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["D14"].value = sum(matched_gstr2a_3way[R2_TOTAL_TAX])
    ws["E14"].value = len(matched_itr_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["F14"].value = sum(matched_itr_3way[PR_TOTAL_TAX])

    bal_gstr2a_1cut = gstr2a_work[~mask_1a]
    bal_itr_1cut = itr_work[~mask_1b]

    # Second Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv No

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_1cut, values=R2_TOTAL_TAX, index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_1cut, values=PR_TOTAL_TAX, index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_2_WAY", right_on="GST_INVNO_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list1 = compared[mask_1]["GST_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a Boolean Array

    mask_1b = bal_itr_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a boolean array

    matched_gstr2a_2way1 = bal_gstr2a_1cut[mask_1a]
    matched_itr_2way1 = bal_itr_1cut[mask_1b]

    matched_gstr2a_2way1["Matching Category"] = "2 Way matching GST + Inv No"
    matched_gstr2a_2way1["Final_Effcorp _Remarks"] = "Exact Matches with Amount- GSTIN 2Way Inv No"
    
    
    matched_itr_2way1["Matching Category"] = "2 Way matching GST + Inv No"
    matched_itr_2way1["Final_Effcorp _Remarks"] = "Exact Matches with Amount- GSTIN 2Way Inv No"

    
    ws["C15"].value = len(matched_gstr2a_2way1["GST_INVNO_2_WAY"])
    ws["D15"].value = sum(matched_gstr2a_2way1[R2_TOTAL_TAX])
    ws["E15"].value = len(matched_itr_2way1["GST_INVNO_2_WAY"])
    ws["F15"].value = sum(matched_itr_2way1[PR_TOTAL_TAX])

    bal_gstr2a_2cut = bal_gstr2a_1cut[~mask_1a]
    bal_itr_2cut = bal_itr_1cut[~mask_1b]

    # Third Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_2cut, values=R2_TOTAL_TAX, index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_2cut, values=PR_TOTAL_TAX, index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVDATE_2_WAY", right_on="GST_INVDATE_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list2 = compared[mask_1]["GST_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a boolean array

    matched_gstr2a_2way2 = bal_gstr2a_2cut[mask_1a]
    matched_itr_2way2 = bal_itr_2cut[mask_1b]

    matched_gstr2a_2way2["Matching Category"] = "2 Way matching GST + Inv Date"
    matched_gstr2a_2way2["Final_Effcorp _Remarks"] = "Exact Matches with Amount- GSTIN 2Way Inv Date"

    
    matched_itr_2way2["Matching Category"] = "2 Way matching GST + Inv Date"
    matched_itr_2way2["Final_Effcorp _Remarks"] = "Exact Matches with Amount- GSTIN 2Way Inv Date"

    ws["C16"].value = len(matched_gstr2a_2way2["GST_INVDATE_2_WAY"])
    ws["D16"].value = sum(matched_gstr2a_2way2[R2_TOTAL_TAX])
    ws["E16"].value = len(matched_itr_2way2["GST_INVDATE_2_WAY"])
    ws["F16"].value = sum(matched_itr_2way2[PR_TOTAL_TAX])

    bal_gstr2a_3cut = bal_gstr2a_2cut[~mask_1a]
    bal_itr_3cut = bal_itr_2cut[~mask_1b]

    print(f"The 3 way matching using GST is done.... Now ,we are doing the mtching using PAN..Please wait...!!")




    #after the 3 cut matching, now we try to find out the Possible matches in form of PAN matching and upper /lower case matching
    
    # Fourth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_3cut, values=R2_TOTAL_TAX, index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_3cut, values=PR_TOTAL_TAX, index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_INVDATE_3_WAY", right_on="PAN_INVNO_INVDATE_3_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_3_way_list2 = compared[mask_1]["PAN_INVNO_INVDATE_3_WAY"].values

    mask_1a = bal_gstr2a_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a boolean array

    matched_gstr2a_3way2 = bal_gstr2a_3cut[mask_1a]
    matched_itr_3way2 = bal_itr_3cut[mask_1b]

    matched_gstr2a_3way2["Matching Category"] = "3 Way matching PAN + Inv No+ Inv Date"
    matched_gstr2a_3way2["Final_Effcorp _Remarks"] = "Exact Matches with Amount- PAN 3way"
    
    matched_itr_3way2["Matching Category"] = "3 Way matching PAN + Inv No + Inv Date"
    matched_itr_3way2["Final_Effcorp _Remarks"] = "Exact Matches with Amount- PAN 3 way"

    ws["C20"].value = len(matched_gstr2a_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["D20"].value = sum(matched_gstr2a_3way2[R2_TOTAL_TAX])
    ws["E20"].value = len(matched_itr_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["F20"].value = sum(matched_itr_3way2[PR_TOTAL_TAX])

    bal_gstr2a_4cut = bal_gstr2a_3cut[~mask_1a]
    bal_itr_4cut = bal_itr_3cut[~mask_1b]

    # Fifth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_4cut, values=R2_TOTAL_TAX, index=["PAN_INVNO_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values=PR_TOTAL_TAX, index=["PAN_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_2_WAY", right_on="PAN_INVNO_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list3 = compared[mask_1]["PAN_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a Boolean Array

    mask_1b = bal_itr_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a boolean array

    matched_gstr2a_2way3 = bal_gstr2a_4cut[mask_1a]
    matched_itr_2way3 = bal_itr_4cut[mask_1b]

    matched_gstr2a_2way3["Matching Category"] = "2 Way matching PAN + Inv No"
    matched_gstr2a_2way3["Final_Effcorp _Remarks"] = "Exact Matches with Amount- PAN 2way Inv No"
    
    matched_itr_2way3["Matching Category"] = "2 Way matching PAN + Inv No "
    matched_itr_2way3["Final_Effcorp _Remarks"] = "Exact Matches with Amount- PAN 2way Inv No"

    ws["C21"].value = len(matched_gstr2a_2way3["PAN_INVNO_2_WAY"])
    ws["D21"].value = sum(matched_gstr2a_2way3[R2_TOTAL_TAX])
    ws["E21"].value = len(matched_itr_2way3["PAN_INVNO_2_WAY"])
    ws["F21"].value = sum(matched_itr_2way3[PR_TOTAL_TAX])

    bal_gstr2a_5cut = bal_gstr2a_4cut[~mask_1a]
    bal_itr_5cut = bal_itr_4cut[~mask_1b]



    # Sixth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_5cut, values=R2_TOTAL_TAX, index=["PAN_INVDATE_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values=PR_TOTAL_TAX, index=["PAN_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVDATE_2_WAY", right_on="PAN_INVDATE_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list4 = compared[mask_1]["PAN_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a Boolean Array

    mask_1b = bal_itr_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a boolean array

    matched_gstr2a_2way4 = bal_gstr2a_5cut[mask_1a]
    matched_itr_2way4 = bal_itr_5cut[mask_1b]

    matched_gstr2a_2way4["Matching Category"] = "2 Way matching PAN + Inv Date"
    matched_gstr2a_2way4["Final_Effcorp _Remarks"] = "Exact Matches with Amount- PAN 2way Inv Date"
    
    matched_itr_2way4["Matching Category"] = "2 Way matching PAN + Inv Date "
    matched_itr_2way4["Final_Effcorp _Remarks"] = "Exact Matches with Amount- PAN 2way Inv Date"


    ws["C22"].value = len(matched_gstr2a_2way4["PAN_INVDATE_2_WAY"])
    ws["D22"].value = sum(matched_gstr2a_2way4[R2_TOTAL_TAX])
    ws["E22"].value = len(matched_itr_2way4["PAN_INVDATE_2_WAY"])
    ws["F22"].value = sum(matched_itr_2way4[PR_TOTAL_TAX])

    bal_gstr2a_6cut = bal_gstr2a_5cut[~mask_1a]
    bal_itr_6cut = bal_itr_5cut[~mask_1b]

    
    #NOw, after all matching, we are further analyzing the et Unmatched Cases


    print(f"Analyzing the Unmatched cases of ITR.... Please wait..!")

    #First, we check whether the PAN we are searching is present in GSTr2A at all or not. If not present then
    #we identify it separately. These have absolute no chaces of matching


    pan_itr=list(set(list(bal_itr_6cut["PAN_Number"].values)))
    pan_2a=bal_gstr2a_6cut["PAN_Number"].values

    maskpan=bal_itr_6cut["PAN_Number"].isin(pan_2a)

    bal_itr_6cut1=bal_itr_6cut[maskpan]

    unmatched_itr1=bal_itr_6cut[~maskpan]

    unmatched_itr1["Remarks_Effcorp"]="PAN/GST not available in GSTR2A"
    unmatched_itr1["Final_Effcorp _Remarks"] = "Unmatched- PAN/GST not in GSTR2A_2B"

    
    ws["E25"].value = len(unmatched_itr1["Remarks_Effcorp"])
    ws["F25"].value = sum(unmatched_itr1[PR_TOTAL_TAX])


    #Second, we will see the CheckSUm Digit of the GST Number. Whether the last charater which is acheck sum is matching or not
    #this is also very crucials, as If GSTIN is invalid, there is no point of matching



    bal_itr_6cut1["GSTN Status"]=bal_itr_6cut1[PR_VENDOR_GSTREG_NO].apply(lambda x:gstchecksum(x))

    mask1=bal_itr_6cut1["GSTN Status"].values=="Check Sum MATCH"


    bal_itr_6cut2=bal_itr_6cut1[mask1]

    unmatched_itr2=bal_itr_6cut1[~mask1]


    unmatched_itr2["Remarks_Effcorp"]="GST Number Check Sum Incorrect"
    unmatched_itr2["Final_Effcorp _Remarks"] = "Unmatched- CHECK SUM INCORRECT"



    ws["E26"].value = len(unmatched_itr2["Remarks_Effcorp"])
    ws["F26"].value = sum(unmatched_itr2[PR_TOTAL_TAX])


    #Third, we will be checking the Invoice Number check
    #if Invoice Number exceeds 16 digits , then we will be marking these seprately as no chaces of matching

    bal_itr_6cut2["Invoice No Check"]=bal_itr_6cut2[PR_INVOICE_NUMBER].apply(lambda x:gstinvcheck(x))

    mask2=bal_itr_6cut2["Invoice No Check"].values=="Invoice Number Valid"


    bal_itr_6cut3=bal_itr_6cut2[mask2]
    unmatched_itr3=bal_itr_6cut2[~mask2]

    unmatched_itr3["Remarks_Effcorp"]="Invoice No length exceed 16 digit"
    unmatched_itr3["Final_Effcorp _Remarks"] = "Unmatched- INV LENGTH EXCEED"

    

    ws["E27"].value = len(unmatched_itr3["Remarks_Effcorp"])
    ws["F27"].value = sum(unmatched_itr3[PR_TOTAL_TAX])


    
    #Here, Before we do the Fuzzy Logic, I would want to do the 3 way match using PAN , and for the cases other than Within the 
    #tolerance limit, we will say as probable matches, where only the amount is different
    
    #Other than amount matching but PAN 3 way matching
    
    gstr2a_pivot = pd.pivot_table(bal_gstr2a_6cut, values=R2_TOTAL_TAX, index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_6cut3, values=PR_TOTAL_TAX, index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={R2_TOTAL_TAX: 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={PR_TOTAL_TAX: 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_INVDATE_3_WAY", right_on="PAN_INVNO_INVDATE_3_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    
    #So,here we have increased the toleraance limit by 100 times.
    #so, if tol liit was 30,then difference upto 3000 will be considered as a Possible match
    
    conditions = [compared["Difference_in_Tax"] > (tol_limit*100),

                  compared["Difference_in_Tax"] < (tol_limit * -100),
                  
                  ((compared["Tax_as_per_ITR"] < 1) & (compared["Tax_as_per_ITR"] >= -1)),
                  
                  (((compared["Tax_as_per_ITR"]>1) | (compared["Tax_as_per_ITR"]<-1)) & ((compared["Difference_in_Tax"] > ((tol_limit+1) * -100)) & (compared["Difference_in_Tax"] < ((tol_limit+1)*100))))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",
               
               "Not Present in ITR",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)
    
    
    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_3_way_list2 = compared[mask_1]["PAN_INVNO_INVDATE_3_WAY"].values

    mask_1a = bal_gstr2a_6cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_6cut3["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a boolean array

    prob_gstr2a_match2 = bal_gstr2a_6cut[mask_1a]
    prob_itr_match2 = bal_itr_6cut3[mask_1b]

    prob_gstr2a_match2["Matching Category"] = "Probable 3 Way matching with PAN- Amount Difference"
    prob_gstr2a_match2["Final_Effcorp _Remarks"] = "Probable Matches with Amount Diff- PAN 3way"
    
    prob_itr_match2["Matching Category"] = "Probable 3 Way matching with PAN- Amount Difference"
    prob_itr_match2["Final_Effcorp _Remarks"] = "Probable Matches with Amount Diff- PAN 3way"

    ws["C17"].value = len(prob_gstr2a_match2["PAN_INVNO_INVDATE_3_WAY"])
    ws["D17"].value = sum(prob_gstr2a_match2[R2_TOTAL_TAX])
    ws["E17"].value = len(prob_itr_match2["PAN_INVNO_INVDATE_3_WAY"])
    ws["F17"].value = sum(prob_itr_match2[PR_TOTAL_TAX])

    bal_gstr2a_6cut2 = bal_gstr2a_6cut[~mask_1a]
    bal_itr_6cut4 = bal_itr_6cut3[~mask_1b]
    
    
    
    
    
    
    
    
    #hre we will ttry to do the fuzzy matching of the Invoice Number with the GSTR2A

    print("Trying to do some Fuzzy matches in GSTR2A and ITR. Please wait....!!")


    from difflib import SequenceMatcher , get_close_matches

    cant_match=[]
    matches_itr=[]
    matches_gstr2ai=[]

    df=list(set(list(bal_itr_6cut4["PAN_Number"].values)))

    for i in df:
    #     Here the i variable is storing the PAN number each time the loop runs
        itr_balinv=bal_itr_6cut4[bal_itr_6cut4["PAN_Number"].values==i][PR_INVOICE_NUMBER].values
        
        # print(f"This is ITR Invoice of {i}")
        # print(itr_balinv)
        
        gstr2a_balinv=(bal_gstr2a_6cut2[bal_gstr2a_6cut2["PAN_Number"].values==i]["Inv_CN_DN_Number_Finall"].values).tolist()
        # print(f"This is GSTR2A Invoice of {i}")
        # print(gstr2a_balinv)
        

        
        zipped=zip(itr_balinv,gstr2a_balinv)
        
        
        if len(gstr2a_balinv)==0:
            cant_match.append(itr_balinv)
            
        
            
        else:
               
            for inv in itr_balinv:
                
                matches_gstr2a=get_close_matches(inv,gstr2a_balinv,n=1,cutoff=0.90)
                
                if len(matches_gstr2a)==1:
                    
                    # print(f"this is inv{inv}")
                    matches_itr.append(inv)
                    matches_gstr2ai.append(matches_gstr2a[0])
                    
                    try:
                        gstr2a_balinv.remove(matches_gstr2a[0])
                    except:
                        continue
                else:
                    
                    continue
            
          
            
            cant_match.append(list(set(itr_balinv)-set(matches_itr)))
        

    mask1a=bal_itr_6cut4[PR_INVOICE_NUMBER].isin(matches_itr)

    mask1b=bal_gstr2a_6cut2["Inv_CN_DN_Number_Finall"].isin(matches_gstr2ai)



    prob_itr_match=bal_itr_6cut4[mask1a]

    prob_gstr2a_match=bal_gstr2a_6cut2[mask1b]

    prob_gstr2a_match["Matching Category"] = "Probable Match- Fuzzy Logic"
    prob_gstr2a_match["Final_Effcorp _Remarks"] = "Prob Matches with PAN _Invoice No"
    
    prob_itr_match["Matching Category"] = "Probable Match- Fuzzy Logic"
    prob_itr_match["Final_Effcorp _Remarks"] = "Prob Matches with PAN _Invoice No"



    bal_itr_6cut5=bal_itr_6cut4[~mask1a]

    bal_gstr2a_7cut=bal_gstr2a_6cut2[~mask1b]

    bal_itr_6cut5["Remarks_Effcorp"]="These Cases are Not Matching"
    bal_itr_6cut5["Final_Effcorp _Remarks"] = "Finally Not Matching"

    bal_gstr2a_7cut["Final_Effcorp _Remarks"] = "Finally Not Matching"


    ws["C18"].value = len(prob_gstr2a_match["Inv_CN_DN_Number_Finall"])
    ws["D18"].value = sum(prob_gstr2a_match[R2_TOTAL_TAX])
    ws["E18"].value = len(prob_itr_match[PR_INVOICE_NUMBER])
    ws["F18"].value = sum(prob_itr_match[PR_TOTAL_TAX])


    print(f"Matchig is done...Creating the 2 Files for you. Summary.xlsx & Working.xlsx")



    #now, we will be merging all these Unmatched cases of itr and final balance cut ITR

    bal_itr_7cut=pd.concat([unmatched_itr1,unmatched_itr2,unmatched_itr3,bal_itr_6cut5])


    combo_gstr2a = pd.concat([gstr2a_not_filed,gstr2a_rcm,gstr2a_notax,matched_gstr2a_3way, matched_gstr2a_2way1, 
                              matched_gstr2a_2way2,matched_gstr2a_3way2, matched_gstr2a_2way3,matched_gstr2a_2way4,
                              prob_gstr2a_match,prob_gstr2a_match2,bal_gstr2a_7cut], ignore_index=True)

    combo_itr = pd.concat([matched_itr_3way, matched_itr_2way1, matched_itr_2way2,matched_itr_3way2,
                           matched_itr_2way3, matched_itr_2way4,prob_itr_match,prob_itr_match2,bal_itr_7cut], ignore_index=True)
    
    combo_gstr2a.to_excel(writer, sheet_name='Orignal GSTR2A', index=False)

    combo_itr.to_excel(writer, sheet_name='Original ITR', index=False)

    all_matched_2a = pd.concat([matched_gstr2a_3way, matched_gstr2a_2way1, matched_gstr2a_2way2,matched_gstr2a_3way2, matched_gstr2a_2way3,matched_gstr2a_2way4,prob_gstr2a_match,prob_gstr2a_match2], ignore_index=True)

    all_matched_itr = pd.concat([matched_itr_3way, matched_itr_2way1, matched_itr_2way2,matched_itr_3way2,matched_itr_2way3, matched_itr_2way4,prob_itr_match,prob_itr_match2], ignore_index=True)

    all_matched_2a.to_excel(writer, sheet_name='Matched_GSTR2A', index=False)

    all_matched_itr.to_excel(writer, sheet_name='Matched_ITR', index=False)

    bal_gstr2a_7cut.to_excel(writer, sheet_name='Unmatched_GSTR2A', index=False)

    bal_itr_7cut.to_excel(writer, sheet_name='Unmatched_ITR', index=False)


    ws["C28"].value = len(bal_gstr2a_7cut["GST_INVDATE_2_WAY"])
    ws["D28"].value = sum(bal_gstr2a_7cut[R2_TOTAL_TAX])
    ws["E28"].value = len(bal_itr_6cut5["Remarks_Effcorp"])
    ws["F28"].value = sum(bal_itr_6cut5[PR_TOTAL_TAX])



    writer.close()

    print("Success! ")



    wb.save(fullpath2)
    writer.close()

    wb.close()
    writer.close()


    print(f'Matching has been done and saved in below path \n {fullpath2}\n{fullpath1} ')

    # return (writer)
    return {
        "writer": writer,
        "working": fullpath1,
        "summary": fullpath2,
    }






def get_gst_type(item):

    """

    This function will identify the type of the GST Number .

    It will classify whether a particular GSTIN is TDS related, TCE Related, OIDAR, UN , Govt Dept, or Normal ISD COmposition related.


    Also, in case the GSTIN is not valid, the same shall also be returned as Invalid GSTN

    """

    import re
    

    unbody = re.compile("[0-9]{4}[A-Z]{3}[0-9]{5}[UO]{1}[N][A-Z0-9]{1}")

    govt_depid = re.compile("[0-9]{2}[a-zA-Z]{4}[0-9]{5}[a-zA-Z]{1}[0-9]{1}[Z]{1}[0-9]{1}")

    nri_id = re.compile("[0-9]{4}[a-zA-Z]{3}[0-9]{5}[N][R][0-9a-zA-Z]{1}")

    tds_id = re.compile("[0-9]{2}[a-zA-Z]{4}[a-zA-Z0-9]{1}[0-9]{4}[a-zA-Z]{1}[1-9A-Za-z]{1}[D]{1}[0-9a-zA-Z]{1}")

    tcs_id = re.compile("[0-9]{2}[a-zA-Z]{5}[0-9]{4}[a-zA-Z]{1}[1-9A-Za-z]{1}[C]{1}[0-9a-zA-Z]{1}")

    oidar_id = re.compile("[9][9][0-9]{2}[a-zA-Z]{3}[0-9]{5}[O][S][0-9a-zA-Z]{1}")

    
    norm_com_isd = re.compile("[0-9]{2}[a-zA-Z]{5}[0-9]{4}[a-zA-Z]{1}[1-9A-Za-z]{1}[Zz1-9A-Ja-j]{1}[0-9a-zA-Z]{1}")

    if item[-1]==getgstcheck(item):

        if oidar_id.search(item):
            gstn_type = "OIDAR ID GSTN"
            return (gstn_type)

        elif unbody.search(item):
            gstn_type = "UN BODY GSTN"
            return (gstn_type)
        elif govt_depid.search(item):
            gstn_type = "GOVT DEPT ID GSTN"
            return (gstn_type)
        elif nri_id.search(item):
            gstn_type = "NRI GSTN"
            return (gstn_type)
        elif tds_id.search(item):
            gstn_type = "TDS ID GSTN"
            return (gstn_type)
        elif tcs_id.search(item):
            gstn_type = "TCS ID GSTN"
            return (gstn_type)
        elif oidar_id.search(item):
            gstn_type = "OIDAR ID GSTN"
            return (gstn_type)
        elif norm_com_isd.search(item):
            gstn_type = "Normal_Composition_ISD GSTIN"
            return (gstn_type)

        else:
            gstn_type="Could not verify GSTN"
            return (gstn_type)


    else:
        gstn_type=("Invalid GSTN")
        return (gstn_type)





